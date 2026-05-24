# 整个流程
# （用户）通过一个按钮，打开选择文件窗口，选择了xlsx文件 需求，无参函数
# （函数）打开选择文件窗口，拿到选择的文件地址
# （函数）加载文件，读取文本表清单
# （用户）选择需要使用的工作表
# （函数）读取所有图片内容
# （函数）对所有图片，进行通用OCR识别，根据识别结果使用银行卡OCR和身份证OCR进行二次识别，拿到精准结果
# （函数）在表格每条记录的右侧，添加相关的数据
# （用户）得到完成消息

import logging
import tkinter as tk
from tkinter import filedialog, ttk
import json
from pathlib import Path
import openpyxl
import openpyxl.worksheet
import base64
from baidu_ocr import general_ocr, get_cached_access_token
from extractor import ExcelImageExtractor

# 配置项
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)

OUTPUT_PATH = "extracted_images"
_MEMORY_FILE = Path.home() / ".file_selector_memory.json"
_DEFAULT_DIR = Path.home() / "Documents"


def _load_last_directory() -> Path:
    """从文件中读取上次访问的目录位置"""
    if not _MEMORY_FILE.exists():
        return None
    try:
        with open(_MEMORY_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        last_dir = data.get("last_directory")
        if last_dir:
            path = Path(last_dir)
            if path.is_dir():
                return path
    except (json.JSONDecodeError, OSError, KeyError) as e:
        logger.warning(f"读取记忆文件失败：{e}")


def _save_last_directory(directory: Path) -> None:
    """保存用户选择目录到记忆文件"""
    try:
        with open(_MEMORY_FILE, "w", encoding="utf-8") as f:
            json.dump({"last_directory": str(directory)}, f, indent=2)
    except OSError as e:
        logger.warning(f"保存记忆文件失败：{e}")


def _openwindow() -> Path:
    """打开文件选择窗口，返回所选文件路径；用户取消返回空字符串；不允许选择多个文件"""
    initial_dir = _load_last_directory()
    if not initial_dir or not initial_dir.exists():
        initial_dir = _DEFAULT_DIR if _DEFAULT_DIR.exists() else Path.home()

    # 创建根窗口并隐藏
    root = tk.Tk()
    root.withdraw()

    # 弹出文件选择对话框
    file_path = filedialog.askopenfilename(
        title="选择文件",
        initialdir=str(initial_dir),
        filetypes=[("xlsx文件", "*.xlsx")],
    )

    root.destroy()
    if file_path:
        selected = Path(file_path)
        _save_last_directory(selected.parent)
        return selected
    return Path()


def _select_sheet(wb: openpyxl.Workbook):
    """弹出选择工作表对话框，返回选中的工作表名"""
    sheet_names = wb.sheetnames
    if not sheet_names:
        print("工作簿中无工作表")
        return None

    win = tk.Toplevel()
    win.title("选择工作表")
    win.geometry("300x150")
    win.resizable(False, False)

    tk.Label(win, text="请选择要使用的工作表").pack(pady=10)
    combo = ttk.Combobox(win, values=sheet_names, state="readonly")
    combo.pack(pady=5)
    combo.current(0)

    result = [None]

    def on_ok():
        result[0] = combo.get()
        win.destroy()

    def on_cancel():
        win.destroy()

    tk.Button(win, text="确定", command=on_ok).pack(side=tk.LEFT, padx=20, pady=10)
    tk.Button(win, text="取消", command=on_cancel).pack(side=tk.RIGHT, padx=20, pady=10)

    win.grab_set()
    win.wait_window()
    return result[0]


def _image2base64(img_bytes) -> str:
    img_base64 = base64.b64encode(img_bytes).decode("utf-8")
    return img_base64


def _find_data_start_row(sheet):
    """找到第一条数据行（A列为数字且不为空的行）"""
    for row in range(1, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=1).value
        if (
            cell_value is not None
            and isinstance(cell_value, (int, float))
            and cell_value != ""
        ):
            return row
    return None


def excel_process():
    file_path = _openwindow()
    if not file_path:
        print("未选择文件")
        return
    wb = openpyxl.load_workbook(filename=file_path)
    selected_sheet = _select_sheet(wb)
    if selected_sheet:
        sheet = wb[selected_sheet]
    else:
        print("未选择工作表")
        return

    data_start_row = _find_data_start_row(sheet)
    if data_start_row is None:
        print("未找到数据行")
        return

    header_row = data_start_row - 1

    max_col = sheet.max_column
    new_id_col = max_col + 1
    new_bankcard_col = max_col + 2
    new_bankname_col = max_col + 3
    sheet.cell(row=header_row, column=new_id_col, value="身份证号")
    sheet.cell(row=header_row, column=new_bankcard_col, value="银行卡号")
    sheet.cell(row=header_row, column=new_bankname_col, value="开户行")

    # images = _getimage(file_path,sheet)
    e_extor = ExcelImageExtractor(file_path)
    images = e_extor.get_float_images(selected_sheet)
    if not images:
        logger.info("当前工作表中没有图片")
        return

    try:
        access_token = get_cached_access_token()
    except Exception as e:
        logger.error(f"无法获得access_token:{e}")
        return

    for row, img_bytes in images:
        # raw_data = my_img.get_data()
        img_base64 = _image2base64(img_bytes)
        # print('目前执行到此为止')
        # exit()
        result = general_ocr(img_base64, access_token)
        if not result:
            logger.error("OCR无结果返回")
        else:
            if result["type"] == "idcard":
                sheet.cell(row=row, column=new_id_col, value=result["data"][1])
            if result["type"] == "bankcard":
                sheet.cell(row=row, column=new_bankcard_col, value=result["data"][0])
                sheet.cell(row=row, column=new_bankname_col, value=result["data"][3])

    output_path = file_path.parent / f"{file_path.stem}_processed.xlsx"
    wb.save(output_path)
    logger.info("完成", f"处理完成！\n结果已保存至：{output_path}")


def excel_process2(xlsx_path: Path, sheetname: str):
    if not xlsx_path.exists():
        logger.error(f"文件不存在:{xlsx_path}")
        return
    try:
        wb = openpyxl.load_workbook(filename=xlsx_path)
        if sheetname not in wb.sheetnames:
            logger.error(f"工作表{sheetname}不在文件中")
            return
        sheet = wb[sheetname]
        output_path = xlsx_path.parent / f"{xlsx_path.stem}_processed.xlsx"
    except Exception as e:
        logger.error(f"加载文件失败{e}")
        return
    data_start_row = _find_data_start_row(sheet)
    if data_start_row is None:
        logger.error("未找到数据行")
        return
    header_row = data_start_row - 1
    max_col = sheet.max_column
    new_id_col = max_col + 1
    new_bankcard_col = max_col + 2
    new_bankname_col = max_col + 3
    sheet.cell(row=header_row, column=new_id_col, value="身份证号")
    sheet.cell(row=header_row, column=new_bankcard_col, value="银行卡号")
    sheet.cell(row=header_row, column=new_bankname_col, value="开户行")
    extractor = ExcelImageExtractor(xlsx_path)
    images = extractor.get_float_images(sheet)
    if not images:
        logger.info("当前表中没有图片")
        return

    try:
        access_token = get_cached_access_token()
    except Exception as e:
        logger.error(f"无法获得access_token:{e}")
        return

    for row, img_bytes in images:
        img_base64 = _image2base64(img_bytes)
        result = general_ocr(img_base64, access_token)
        if not result:
            logger.error(f"第{row}行图片 OCR 无返回结果")
            continue
        if result["type"] == "idcard":
            sheet.cell(row=row, column=new_id_col, value=result["data"][1])
        if result["type"] == "bankcard":
            sheet.cell(row=row, column=new_bankcard_col, value=result["data"][0])
            sheet.cell(row=row, column=new_bankname_col, value=result["data"][3])

    try:
        wb.save(output_path)
        logger.info(f"处理完成，结果保存至{output_path}")
    except Exception as e:
        logger.error(f"保存文件失败:{e}")


if __name__ == "__main__":
    # 创建主窗口
    root = tk.Tk()
    root.title("Excel 图片识别工具")
    root.geometry("400x200")
    root.resizable(False, False)

    # 说明标签
    label = tk.Label(
        root,
        text="点击按钮选择 Excel 文件，\n程序将自动识别其中的身份证/银行卡图片\n并将结果写入新文件。",
        justify="left",
        pady=20,
    )
    label.pack()

    # 按钮直接绑定无参函数 excel_process()
    btn = tk.Button(
        root,
        text="开始处理",
        command=excel_process,  # 注意：excel_process 必须定义为无参数函数
        width=20,
        height=2,
        bg="#4CAF50",
        fg="white",
    )
    btn.pack(pady=20)

    # 运行主循环
    root.mainloop()
